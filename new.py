import openpyxl
import itertools
import numpy as np
from datetime import datetime

COMBOS = ["VA30", "VA31"]
WAREHOUSE_IDS = ["NY", "CA", "TX"]
IGNORED_CARRIERS = ["Fedex", "Ups"]
MAX_DELAY = 2


def input_warehouses():
    print("Choose Warehouses")

    for i, warehouse in enumerate(WAREHOUSE_IDS):
        print(f"({i + 1}) {warehouse}") #

    print(f"({len(WAREHOUSE_IDS) + 1}) All") #
    print("ex: '1 3'") #

    choices = [int(choice) - 1 for choice in input("> ").split()]
    warehouses = set()

    for choice in choices:
        if choice == len(WAREHOUSE_IDS):
            warehouses.update(range(len(WAREHOUSE_IDS)))
        elif choice in range(len(WAREHOUSE_IDS)):
            warehouses.add(choice)

    return [WAREHOUSE_IDS[wh - 1] for wh in warehouses]


def load_class_lookup(path):
    def parse_lookup_item_nums(item_nums_raw):
        if item_nums_raw is None:
            return []

        try:
            item_nums_raw = item_nums_raw[:item_nums_raw.index("/")]
        except ValueError:
            pass
        return item_nums_raw.split(":")

    class_lookup_sheet = openpyxl.load_workbook(path).active
    class_lookup = {}

    for row in range(2, class_lookup_sheet.max_row + 1):
        item_num_raw = class_lookup_sheet.cell(row, 2).value
        class_name = class_lookup_sheet.cell(row, 1).value

        for item_num in parse_lookup_item_nums(item_num_raw):
            class_lookup[item_num] = class_name

    return class_lookup


def parse_data(data_sheet, class_lookup, warehouses):
    def determine_ship_status(order_time, status):
        if order_time is None:
            return "On Time"

        date_fmt = "%Y-%m-%d"
        current_time = datetime.now()
        business_days = np.busday_count(order_time.strftime(date_fmt), current_time.strftime(date_fmt))
        return "Late" if business_days > MAX_DELAY and status.lower() != "shipped" else "On Time"

    def parse_item_num(item_num):
        if not item_num.startswith("VA"):
            raise ValueError

        return item_num

    def combine_item_nums(item_nums, qtys):
        if len(item_nums) == 1:
            return item_nums[0]

        colors = [item[item.index("-")+1:] if item[item.index("-")+1:] else "" for item in item_nums]

        if any(color != colors[0] for color in colors):
            raise ValueError

        item_nums = [item[:item.index("-")] for item in item_nums]
        num = sum(int(item[-2:]) * qty for item, qty in zip(item_nums, qtys))
        first = max(item_nums, key=lambda v: int(v[2:]))
        return f"{first}-{num}{colors[0]}"

    #================================#

    output_data = {}

    for row in range(2, data_sheet.max_row + 1):
        items = []
        qtys = []

        po = data_sheet.cell(row, 1).value
        carrier = data_sheet.cell(row, 6).value
        status = data_sheet.cell(row, 7).value
        warehouse = data_sheet.cell(row, 8).value
        order_time = data_sheet.cell(row, 3).value
        class_name = ""

        if None in (status, warehouse) or warehouse not in warehouses or carrier in IGNORED_CARRIERS:
            continue

        ship_status = determine_ship_status(order_time, status)
        skip_row = False

        for i in range(data_sheet.max_column + 1):
            item_num_raw = data_sheet.cell(row, 11 + i * 3).value

            if item_num_raw is None:
                break

            try:
                item_num = parse_item_num(item_num_raw)
            except ValueError:
                skip_row = True
                break

            qty = data_sheet.cell(row, 12 + i * 3).value

            if None not in (item_num, qty):
                items.append(item_num)
                qtys.append(int(qty))
                item_class = class_lookup.get(item_num)
                class_name = item_class or ""

        if skip_row:
            continue

        is_combo = all(any(item.startswith(combo) for combo in COMBOS) for item in items)

        total_qty = sum(qtys)

        try:
            uid = combine_item_nums(items, qtys)
        except ValueError:
            num_late = 0 if ship_status != "Late" else total_qty
            nonconflicting = []
            new_qtys = []

            for q, item in zip(qtys, items):
                if item in output_data:
                    num_late = 0 if ship_status != "Late" else q
                    if output_data[item]["meta"]["mode"] != "NORMAL":
                        raise ValueError("Item number merge conflicted with row merge while resolving merge issue")
                    if not is_combo:
                        output_data[uid]["meta"]["qty"] += q
                        output_data[uid]["meta"]["late"] += num_late

                    new_qty = output_data[uid]["meta"]["qty"]
                    new_late = output_data[uid]["meta"]["late"]

                    if ship_status == "Late":
                        output_data[uid]["columns"][0] = [class_name]
                        output_data[uid]["columns"][1] = [uid]
                        output_data[uid]["columns"][2] = [f"{new_qty} ({new_late} Late)"]
                        output_data[uid]["columns"][3].append(ship_status)
                        output_data[uid]["columns"][4].append(po)
                        output_data[uid]["columns"][5].append(carrier)
                        output_data[uid]["columns"][6].append(warehouse)
                else:
                    nonconflicting.append(item)
                    new_qtys.append(q)


            if not nonconflicting:
                continue

            row_uid = " ".join(nonconflicting)

            for q, item in zip(new_qtys, nonconflicting):
                output_data[item] = { "meta": { "parent": row_uid, "q": q } }

            total_qty = sum(new_qtys)
            num_late = 0 if ship_status != "Late" else total_qty

            output_data[row_uid] = {
                "columns": [
                    [class_name],
                    nonconflicting,
                    [f"{total_qty} ({num_late} Late)"],
                    [ship_status],
                    [po],
                    [carrier],
                    [warehouse],
                ],

                "merge": [1, 3, 4, 5, 6, 7],

                "meta": {
                    "mode": "MULTI",
                },
            }
        else:
            if is_combo:
                total_qty = 1

            num_late = 0 if ship_status != "Late" else total_qty

            if uid in output_data:
                if "parent" in output_data[uid]["meta"]:
                    parent = output_data[uid]["meta"]["parent"]
                    output_data[parent]["columns"][1].remove(uid)
                    if not is_combo: total_qty += output_data[uid]["meta"]["q"]

                    if not output_data[parent]["columns"][1]:
                        del output_data[parent]
                    del output_data[uid]

                    num_late = 0 if ship_status != "Late" else total_qty

                    if ship_status == "Late":
                        output_data[uid] = {
                            "columns": [
                                [class_name],
                                [uid],
                                [f"{total_qty} ({num_late} Late)"],
                                [ship_status],
                                [po],
                                [carrier],
                                [warehouse],
                            ],

                            "merge": [1, 2, 3],

                            "meta": {
                                "mode": "NORMAL",
                                "qty": total_qty,
                                "late": num_late
                            },
                        }
                    else:
                        output_data[uid] = {
                            "columns": [
                                [],
                                [],
                                [],
                                [],
                                [],
                                [],
                                [],
                            ],

                            "merge": [1, 2, 3],

                            "meta": {
                                "mode": "NORMAL",
                                "qty": total_qty,
                                "late": num_late
                            },
                        }
                else:
                    if not is_combo:
                        output_data[uid]["meta"]["qty"] += total_qty
                        output_data[uid]["meta"]["late"] += num_late  #

                    new_qty = output_data[uid]["meta"]["qty"]
                    new_late = output_data[uid]["meta"]["late"]

                    if ship_status == "Late":
                        output_data[uid]["columns"][0] = [class_name]
                        output_data[uid]["columns"][1] = [uid]
                        output_data[uid]["columns"][2] = [f"{new_qty} ({new_late} Late)"]
                        output_data[uid]["columns"][3].append(ship_status)
                        output_data[uid]["columns"][4].append(po)
                        output_data[uid]["columns"][5].append(carrier)
                        output_data[uid]["columns"][6].append(warehouse)
            else:
                if ship_status == "Late":
                    output_data[uid] = {
                        "columns": [
                            [class_name],
                            [uid],
                            [f"{total_qty} ({num_late} Late)"],
                            [ship_status],
                            [po],
                            [carrier],
                            [warehouse],
                        ],

                        "merge": [1, 2, 3],

                        "meta": {
                            "mode": "NORMAL",
                            "qty": total_qty,
                            "late": num_late
                        },
                    }
                else:
                    output_data[uid] = {
                        "columns": [
                            [],
                            [],
                            [],
                            [],
                            [],
                            [],
                            [],
                        ],

                        "merge": [1, 2, 3],

                        "meta": {
                            "mode": "NORMAL",
                            "qty": total_qty,
                            "late": num_late
                        },
                    }

    return output_data


def write_data(output_data, path):
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active

    column_headers = [
        "Class",
        "Item",
        "Total Qty",
        "Ship Status",
        "PO",
        "Carrier",
        "Warehouse",
    ]

    for i, column in enumerate(column_headers):
        output_sheet.cell(1, i + 1).value = column

    row_ptr = 2

    for item in output_data.values():
        start_row = row_ptr

        for offset in itertools.count():
            written = False

            for col, data in enumerate(item["columns"]):
                value = ""

                if offset < len(data):
                    value = data[offset]
                    written = True #

                output_sheet.cell(row_ptr + offset, col + 1).value = value

            if not written:
                break

        row_ptr += offset

        if row_ptr - start_row > 1:
            for col in item["merge"]:
                output_sheet.merge_cells(start_row=start_row, end_row=row_ptr - 1, start_column=col, end_column=col) #
                for row in range(start_row, row_ptr):
                    output_sheet.cell(row, col).alignment = openpyxl.styles.Alignment(vertical="center")

    output_wb.save(path)


def main():
    data_sheet = openpyxl.load_workbook("rp3.xlsx").active
    class_lookup = load_class_lookup("class_lookup.xlsx")
    warehouses = input_warehouses()
    output_data = parse_data(data_sheet, class_lookup, warehouses)

    write_data(output_data, "output.xlsx")


if __name__ == "__main__":
    main()
