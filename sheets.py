"""
Module for reading/writing of google sheets with item data
"""

import openpyxl
import itertools
import numpy as np
from datetime import datetime
from json import dumps

COMBOS = ["VA30", "VA31"]
WAREHOUSE_IDS = ["NY", "CA", "TX"]
IGNORED_CARRIERS = ["Fedex", "Ups"]
MAX_DELAY = 2


def decompose_item_num(item_num):
    """
    Removes color information from an item number
    """

    item_num = item_num.replace("-", "")

    try:
        color_index = next(i for i, c in enumerate(item_num[2:]) if c.isalpha()) + 2
    except StopIteration:
        color_index = len(item_num)

    return item_num[:color_index], item_num[color_index:]


def extract_order_nums(path):
    """
    Extracts order numbers from a google sheet downloaded from the order website
    """

    sheet = openpyxl.load_workbook(path).active
    order_nums = []

    for row in range(2, sheet.max_row + 1):
        order_num = str(sheet.cell(row, 3).value)

        if order_num is None:
            break

        order_nums.append(order_num)

    return order_nums


def strip_color(num):
    """
    Strips color information from an item only if it follows a dash
    """

    dash_i = num.index("-")

    try:
        color_i = next(i for i, c in enumerate(num[dash_i:]) if c.isalpha()) + dash_i
        num = num[:color_i]
    except StopIteration:
        pass

    return num


def input_warehouses():
    """
    Read warehouse selection from stdin
    """

    print("Choose Warehouses")

    for i, warehouse in enumerate(WAREHOUSE_IDS):
        print(f"({i + 1}) {warehouse}")

    print(f"({len(WAREHOUSE_IDS) + 1}) All")
    print("ex: '1 3'")

    choices = [int(choice) - 1 for choice in input("> ").split()]
    warehouses = set()

    for choice in choices:
        if choice == len(WAREHOUSE_IDS):
            warehouses.update(range(len(WAREHOUSE_IDS)))
        elif choice in range(len(WAREHOUSE_IDS)):
            warehouses.add(choice)

    return [WAREHOUSE_IDS[wh] for wh in warehouses]


def load_class_lookup(path):
    """
    Loads a class lookup sheet
    """

    def parse_lookup_item_nums(item_nums_raw):
        """
        Parses a sequence of item numbers from a raw string
        """

        if item_nums_raw is None:
            return []

        try:
            item_nums_raw = item_nums_raw[:item_nums_raw.index("/")]
        except ValueError:
            pass
        return item_nums_raw.split(":")

    class_lookup_sheet = openpyxl.load_workbook(path).active
    class_lookup = {}

    for row in range(2, class_lookup_sheet.max_row + 1):
        item_num_raw = class_lookup_sheet.cell(row, 2).value
        class_name = class_lookup_sheet.cell(row, 1).value

        for item_num in parse_lookup_item_nums(item_num_raw):
            class_lookup[item_num] = class_name

    return class_lookup


def load_combo_lookup(path):
    """
    Loads a combo lookup sheet
    """

    combo_lookup_sheet = openpyxl.load_workbook(path).active
    combo_lookup = {}

    for row in range(4, combo_lookup_sheet.max_row + 1):
        combo = combo_lookup_sheet.cell(row, 1).value

        if combo is None:
            break

        qty1, piece1 = combo_lookup_sheet.cell(row, 2).value.split()
        qty2, piece2 = combo_lookup_sheet.cell(row, 3).value.split()

        combo_pieces = {}
        combo_pieces[piece1] = int(qty1[:-2])
        combo_pieces[piece2] = int(qty2[:-2])
        combo_lookup[combo] = combo_pieces

    return combo_lookup


class Item(object):
    """
    An object representing an individual item
    """

    def __init__(self):
        self.num = None
        self.qty = None
        self.ship_status = None
        self.po = None
        self.carrier = None
        self.warehouse = None

    def __str__(self):
        return f"[ {self.num} | {self.qty} | {self.ship_status} | {self.po} | {self.carrier} | {self.warehouse} ]"


class Entry(object):
    """
    An object representing a group of items (a row in the output sheet)
    """

    def __init__(self):
        self.items = []
        self.special_order = False
        self.uid = None
        self.is_combo = False
        self.total_qty = None
        self.late_qty = None

    def __str__(self):
        item_str = "\n\t".join(map(str, self.items))
        special_order_char = '*' if self.special_order else ' '
        return f"[{special_order_char}] {self.uid}\n\t{item_str}\n"

    def count_qtys(self, late, combo_lookup):
        """
        Compute the total quantity of a combo entry
        """

        # the following piece of logic selects either all late items or all items based on the boolean `late`
        # if its true then only late items are selected, otherwise all items are selected
        items = [item for item in self.items if not (item.ship_status != "Late" and late)]
        counts = {}
        combo = combo_lookup[strip_color(self.uid)]

        for item in items:
            key = item.num

            try:
                key = key[:key.index("-")]
            except ValueError:
                pass

            counts[key] = counts.get(key, 0) + item.qty

        vals = [v // combo[k] for k, v in counts.items()]

        if not vals:
            return 0

        assert(all(val == vals[0] for val in vals))
        return vals[0]

    # maybe combine with count_qtys
    def compute_qtys(self, combo_lookup):
        """
        Figure out the total quantity
        """

        if self.is_combo:
            self.total_qty = self.count_qtys(False, combo_lookup)
            self.late_qty = self.count_qtys(True, combo_lookup)
        else:
            self.total_qty = sum(int(item.qty) for item in self.items)
            self.late_qty = sum(int(item.qty) for item in self.items if item.ship_status == "Late")

    def compute_uid(self, combo_lookup):
        """
        Find a unique identifier for this entry (entries with shared uids are combined into one)
        """

        if len(self.items) == 1:
            self.uid = self.items[0].num
            return

        decomposed = [decompose_item_num(item.num) for item in self.items]
        colors = [d[1] for d in decomposed]

        if not all(any(item.num.startswith(combo) for combo in COMBOS) for item in self.items) or any(color != colors[0] for color in colors):
            self.special_order = True
            self.uid = "SPECIAL ORDER: " + "".join(f"\n\t{item.num} ({item.qty})" for item in self.items)
            return

        raw_nums = [d[0] for d in decomposed]
        num = sum(int(raw_num[-2:]) * int(item.qty) for raw_num, item in zip(raw_nums, self.items))
        first = max(raw_nums, key=lambda v: int(v[2:]))
        self.uid = f"{first}-{num}{colors[0]}"

        if self.uid in combo_lookup:
            self.is_combo = True
        else:
            self.uid = self.items[0].num
            self.is_combo = False

    def add_entry(self, entry):
        """
        Combine an entry into this one
        """

        self.items.extend(entry.items)

    def get_combo_num(self, combo_lookup):
        """
        Returns the display string for combos
        """

        combo = combo_lookup[strip_color(self.uid)]
        items_str = " ".join(f"{k} ({v})" for k, v in combo.items())
        return f"{self.uid}: {items_str}"

    def write_to(self, output_data, class_lookup, combo_lookup):
        """
        Writes this entry to an output data array
        """

        self.compute_qtys(combo_lookup)

        class_name = class_lookup.get(self.items[0].num, "")
        item_num = self.get_combo_num(combo_lookup) if self.is_combo else self.uid

        to_display = [item for item in self.items if item.ship_status == "Late"]

        output_data.append({
            "data": [ # 2d array (array of columns)
                [class_name],
                [item_num],
                [self.total_qty],
                [self.late_qty],
                [item.qty for item in to_display],
                [item.po for item in to_display],
                [item.carrier for item in to_display],
                [item.warehouse for item in to_display],
            ],

            "merge": [ 1, 2, 3, 4], # columns to merge
        })


def get_ship_status(order_time, status):
    """
    Determine the ship status ("Late" or "On Time") based on the order time and status ("shipped" or "not shipped")
    """

    if order_time is None:
        return "On Time"

    date_fmt = "%Y-%m-%d"
    current_time = datetime.now()
    business_days = np.busday_count(order_time.strftime(date_fmt), current_time.strftime(date_fmt))
    return "Late" if business_days > MAX_DELAY and status.lower() != "shipped" else "On Time"


def get_data(data_sheet):
    """
    Extracts item data from a sheet
    """

    data = []

    for row in range(2, data_sheet.max_row + 1):
        po = data_sheet.cell(row, 1).value
        order_time = data_sheet.cell(row, 3).value
        carrier = data_sheet.cell(row, 6).value
        status = data_sheet.cell(row, 7).value or "shipped"
        warehouse = data_sheet.cell(row, 8).value
        ship_status = get_ship_status(order_time, status)

        items = []

        for i in range(data_sheet.max_column + 1):
            num = data_sheet.cell(row, 11 + i * 3).value
            qty = data_sheet.cell(row, 12 + i * 3).value

            if num is None:
                break

            items.append((num, int(qty)))

        data.append((
            po,
            carrier,
            status,
            warehouse,
            ship_status,
            items,
        ))

    return data


def parse_data(data, warehouses, class_lookup, combo_lookup):
    """
    Parse data extracted from a sheet
    """

    entries = {}

    for po, carrier, status, warehouse, ship_status, items in data:
        entry = Entry()

        if warehouse not in warehouses:
            continue

        for num, qty in items:
            item = Item()

            item.num = num
            item.qty = qty
            item.ship_status = ship_status
            item.po = po
            item.carrier = carrier
            item.warehouse = warehouse

            entry.items.append(item)

        entry.compute_uid(combo_lookup)

        if carrier in IGNORED_CARRIERS and (len(entry.items) == 1 or entry.is_combo):
            continue

        if entry.uid in entries:
            entries[entry.uid].add_entry(entry)
        else:
            entries[entry.uid] = entry

    output_data = []

    for entry in entries.values():
        entry.write_to(output_data, class_lookup, combo_lookup)

    return output_data


def write_data(output_data, path):
    """
    Write data to an output sheet
    """

    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active

    column_headers = [
        "Class",
        "Item",
        "Total Qty",
        "Late Qty",
        "Late Qtys",
        "PO",
        "Carrier",
        "Warehouse",
    ]

    for i, column in enumerate(column_headers):
        output_sheet.cell(1, i + 1).value = column

    row_ptr = 2

    for item in output_data:
        start_row = row_ptr

        for offset in itertools.count():
            written = False

            for col, data in enumerate(item["data"]):
                value = ""

                if offset < len(data):
                    value = data[offset]
                    written = True

                output_sheet.cell(row_ptr + offset, col + 1).value = value

            if not written:
                break

        row_ptr += offset

        if row_ptr - start_row > 1:
            for col in item["merge"]:
                output_sheet.merge_cells(start_row=start_row, end_row=row_ptr - 1, start_column=col, end_column=col)
                for row in range(start_row, row_ptr):
                    output_sheet.cell(row, col).alignment = openpyxl.styles.Alignment(vertical="center")

    output_wb.save(path)


def main():
    data_sheet = openpyxl.load_workbook("report.xlsx").active
    class_lookup = load_class_lookup("class_lookup.xlsx")
    combo_lookup = load_combo_lookup("combo_lookup.xlsx")
    warehouses = input_warehouses()
    output_data = parse_data(get_data(data_sheet), warehouses, class_lookup, combo_lookup)

    write_data(output_data, "output.xlsx")


if __name__ == "__main__":
    main()
