import numpy as np
from datetime import datetime
import openpyxl


def extract_item_nums(item_num_raw, single=False):
    if item_num_raw is None:
        return []

    try:
        item_num = item_num_raw[:item_num_raw.index("/")]
        item_num = item_num[:item_num.index("-")]
    except ValueError:
        pass
    return item_num if single else item_num.split(":")


def parse_item_num(item_num):
    if not item_num.startswith("VA"):
        raise ValueError()

    return item_num[:-2], int(item_num[-2:])


def main():
    data_sheet = openpyxl.load_workbook("rp2.xlsx").active
    class_lookup_sheet = openpyxl.load_workbook("iil.xlsx").active
    class_lookup = {}
    output_data = {}
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active

    for col in range(11, data_sheet.max_row + 1):
        if data_sheet.cell(row=1, column=col).value is None:
            max_items = (col - 11) // 3
            break

    for row in range(2, class_lookup_sheet.max_row + 1):
        item_num_raw = class_lookup_sheet.cell(row=row, column=2).value

        for item_num in extract_item_nums(item_num_raw):
            class_lookup[item_num] = class_lookup_sheet.cell(row=row, column=1).value

    for row in range(2, data_sheet.max_row + 1):
        total_qty = 0
        po = data_sheet.cell(row=row, column=1).value
        carrier = data_sheet.cell(row=row, column=6).value
        order_time = data_sheet.cell(row=row, column=3).value
        class_name = "???"
        item_num = None
        total_n = 0

        if None in (po, carrier, order_time) or carrier in ("Ups", "Fedex"):
            continue

        current_time = datetime.now()
        business_days = np.busday_count(order_time.strftime("%Y-%m-%d"), current_time.strftime("%Y-%m-%d"))
        ship_status = "Late" if business_days > 2 else "On Time"

        for i in range(max_items):
            in_num, n = parse_item_num(extract_item_nums(data_sheet.cell(row=row, column=11 + i * 3).value, single=True))
            item_num = item_num or in_num
            total_n += n

            qty = data_sheet.cell(row=row, column=12 + i * 3).value

            if item_num is not None and qty is not None:
                items.append((item_num, n))

                class_n = class_lookup.get(item_num)
                if class_n is not None:
                    class_name = class_n

                total_qty += qty

        output_data[item_num] = [
            class_name,
            total_n,
            total_qty,
            ship_status,
            po,
            carrier,
        ]
        output_data.append([
            class_name,
            items,
            total_qty,
            ship_status,
            po,
            carrier,
        ])

    columns = [
        "Class",
        "Item",
        "Total Qty",
        "Ship Status",
        "PO",
        "Carrier",
    ]

    for i, column_name in enumerate(columns):
        output_sheet.cell(row=1, column=i + 1).value = column_name

    row_ptr = 2

    for class_name, items, total_qty, ship_status, po, carrier in output_data:
        write_common = True
        start_row = row_ptr
        for item in items:
            if write_common:
                write_common = False
                output_sheet.cell(row=row_ptr, column=1).value = class_name
                output_sheet.cell(row=row_ptr, column=3).value = total_qty
                output_sheet.cell(row=row_ptr, column=4).value = ship_status
                output_sheet.cell(row=row_ptr, column=5).value = po
                output_sheet.cell(row=row_ptr, column=6).value = carrier
            output_sheet.cell(row=row_ptr, column=2).value = item
            row_ptr += 1

        if len(items) > 1:
            output_sheet.merge_cells(start_row=start_row, start_column=1, end_row=row_ptr - 1, end_column=1)
            output_sheet.merge_cells(start_row=start_row, start_column=3, end_row=row_ptr - 1, end_column=3)
            output_sheet.merge_cells(start_row=start_row, start_column=4, end_row=row_ptr - 1, end_column=4)
            output_sheet.merge_cells(start_row=start_row, start_column=5, end_row=row_ptr - 1, end_column=5)

    output_wb.save("out.xlsx")


if __name__ == "__main__":
    main()
