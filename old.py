import numpy as np
from datetime import datetime
import openpyxl

DEFAULT_ITEM_CLASS = "???"
COMBOS = ["VA30", "VA31"]


def extract_item_num(item_num):
    if not item_num.startswith("VA"):
        raise ValueError

    return item_num


def extract_item_nums(item_nums):
    if item_nums is None:
        return []

    try:
        item_nums = item_nums[:item_nums.index("/")]
    except ValueError:
        pass
    return item_nums.split(":")


def combine_item_nums(item_nums, item_qtys):
    if len(item_nums) == 1:
        return item_nums[0]

    item_nums = [item[:item.index("-")] for item in item_nums]
    num = sum(int(item[-2:]) * qty for item, qty in zip(item_nums, qtys))
    first = max(item_nums, key=lambda v: int(v[2:]))
    return f"{first}-{num}"


if __name__ == "__main__":
    data_sheet = openpyxl.load_workbook("data_sheet.xlsx").active
    class_lookup_sheet = openpyxl.load_workbook("class_lookup.xlsx").active
    class_lookup = {}
    output_data = {}
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active

    warehouse_ids = ["NY", "CA", "TX"]

    print("Choose Warehouses")
    print("(1) NY")
    print("(2) CA")
    print("(3) TX")
    print("(4) All")
    print("ex: '1 3'")

    choices = input("> ")
    warehouses = set()

    for choice in choices.split():
        choice = int(choice)
        if choice == 4:
            warehouses.update((1, 2, 3))
        elif choice in range(1, 4):
            warehouses.add(choice)

    to_display = [warehouse_ids[wh - 1] for wh in warehouses]

    for col in range(11, data_sheet.max_row + 1):
        if data_sheet.cell(row=1, column=col).value is None:
            max_items = (col - 11) // 3
            break

    for row in range(2, class_lookup_sheet.max_row + 1):
        item_num_raw = class_lookup_sheet.cell(row=row, column=2).value

        for item_num in extract_item_nums(item_num_raw):
            class_lookup[item_num] = class_lookup_sheet.cell(row=row, column=1).value

    for row in range(2, data_sheet.max_row + 1):
        items = []
        qtys = []
        po = data_sheet.cell(row=row, column=1).value
        carrier = data_sheet.cell(row=row, column=6).value
        status = data_sheet.cell(row=row, column=7).value
        wh_id = data_sheet.cell(row=row, column=8).value
        order_time = data_sheet.cell(row=row, column=3).value
        class_name = DEFAULT_ITEM_CLASS

        if None in (po, carrier, order_time) or carrier in ("Fedex", "Ups") or wh_id not in to_display:
            continue

        current_time = datetime.now()
        business_days = np.busday_count(order_time.strftime("%Y-%m-%d"), current_time.strftime("%Y-%m-%d"))
        ship_status = "Late" if business_days > 2 and status.lower() != "shipped" else "On Time"

        try:
            for i in range(max_items):
                item_num_raw = data_sheet.cell(row=row, column=11 + i * 3).value
                if item_num_raw is None:
                    break
                item_num = extract_item_num(item_num_raw)
                qty = data_sheet.cell(row=row, column=12 + i * 3).value

                if item_num is not None and qty is not None:
                    items.append(item_num)

                    item_class = class_lookup.get(item_num)
                    if item_class is not None:
                        class_name = item_class

                    qtys.append(int(qty))
        except ValueError:
            continue

        is_combo = all(any(item.startswith(combo) for combo in COMBOS) for item in items)

        uid = combine_item_nums(items, qtys)
        total_qty = sum(qtys)

        if is_combo:
            total_qty = 1

        if uid in output_data:
            if not is_combo: output_data[uid][1] += total_qty
            output_data[uid][2].append(wh_id)
            output_data[uid][3].append(ship_status)
            output_data[uid][4].append(po)
            output_data[uid][5].append(carrier)
        else:
            output_data[uid] = [
                class_name,
                total_qty,
                [wh_id],
                [ship_status],
                [po],
                [carrier],
            ]

    columns = [
        "Class",
        "Item",
        "Total Qty",
        "Ship Status",
        "PO",
        "Carrier",
        "Warehouse",
    ]

    for i, column_name in enumerate(columns):
        output_sheet.cell(row=1, column=i + 1).value = column_name

    row_ptr = 2

    for uid, (class_name, qty, wh_ids, ship_statuses, pos, carriers) in output_data.items():
        start_row = row_ptr
        num_late = sum(ship_status == "Late" for ship_status in ship_statuses)

        if all(wh not in to_display for wh in wh_ids):
            continue

        output_sheet.cell(row=row_ptr, column=1).value = class_name
        output_sheet.cell(row=row_ptr, column=2).value = uid
        output_sheet.cell(row=row_ptr, column=3).value = f"{qty} ({num_late} Late)"
        output_sheet.cell(row=row_ptr, column=1).alignment = openpyxl.styles.Alignment(vertical="center")
        output_sheet.cell(row=row_ptr, column=2).alignment = openpyxl.styles.Alignment(vertical="center")
        output_sheet.cell(row=row_ptr, column=3).alignment = openpyxl.styles.Alignment(vertical="center")

        for wh_id, ship_status, po, carrier in zip(wh_ids, ship_statuses, pos, carriers):
            if wh_id in to_display:
                output_sheet.cell(row=row_ptr, column=4).value = ship_status
                output_sheet.cell(row=row_ptr, column=5).value = po
                output_sheet.cell(row=row_ptr, column=6).value = carrier
                output_sheet.cell(row=row_ptr, column=7).value = wh_id
                row_ptr += 1

        if row_ptr - start_row > 1:
            output_sheet.merge_cells(start_row=start_row, end_row=row_ptr - 1, start_column=1, end_column=1)
            output_sheet.merge_cells(start_row=start_row, end_row=row_ptr - 1, start_column=2, end_column=2)
            output_sheet.merge_cells(start_row=start_row, end_row=row_ptr - 1, start_column=3, end_column=3)

    output_wb.save("oldout.xlsx")
